
import os
from PIL import Image
import torch
from torchvision.transforms import ToTensor
from resmem import transformer, ResMem
from openpyxl import Workbook

# Initialize the ResMem model
model = ResMem(pretrained=True)
# Set the model to inference mode.
model.eval() 

# Specify the directory containing the images
image_dir = 'faces'
#'path/to/your/image/directory'

# Get a list of all image file names in the directory
image_files = [os.path.join(image_dir, f) for f in os.listdir(image_dir) if f.endswith('.jpg')or f.endswith('.JPG') or f.endswith('.png')]


# Create a new workbook
wb = Workbook()

# Create a worksheet in the workbook
ws = wb.active

# Set the initial row number
row_num = 1

# Preprocess each image and stack them to form a batch
#preprocessed_images = []
for image_file in image_files:
    img = Image.open(image_file).convert('RGB')
    file_name = os.path.basename(image_file)
   # print(file_name)
    img_x = transformer(img)
    predictions = model(img_x.view(-1, 3, 227, 227))
    #print(predictions)

    prediction_str = str(predictions.detach().numpy())
    
    # Write the file name and predictions to the Excel file
    ws.cell(row=row_num, column=1, value=file_name)
    ws.cell(row=row_num, column=2, value=prediction_str)
    
    # Move to the next row
    row_num += 1

# Save the workbook to a file
wb.save('output_faces_resmem.xlsx')
    #preprocessed_images.append(img)

#batch = torch.stack(preprocessed_images)

# Make predictions on the batch
#predictions = model(batch.view(-1, 3, 227, 227))

#print(predictions)
